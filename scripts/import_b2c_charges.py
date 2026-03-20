import os
import sys
import openpyxl
import pymysql
import re
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import DB_CONFIG, db_config

XLSX_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'data', '财务账单', '海外仓账单（各自建立不同表-混合文件）', '1510'
)

DB_NAME = db_config.database
TABLE_NAME = 'b2c_order_charges'
SHEET_NAME = 'B2C订单费用B2C Order Charges'


def create_database(connection):
    """创建数据库"""
    cursor = connection.cursor()
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
    cursor.execute(f"USE {DB_NAME}")
    cursor.close()
    print(f"数据库 {DB_NAME} 已创建/已选中")


def read_xlsx_sheet(filepath, sheet_name):
    """读取xlsx文件的指定sheet"""
    data_rows = []
    header = None

    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        print(f"  未找到sheet: {sheet_name}")
        wb.close()
        return header, data_rows

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx == 1:
            header = [str(cell) if cell is not None else '' for cell in row]
        else:
            if any(cell is not None and str(cell).strip() for cell in row):
                data_rows.append([cell for cell in row])

    wb.close()
    return header, data_rows


def sanitize_column_name(name):
    """清理列名，使其符合SQL规范"""
    if not name or not str(name).strip():
        return 'column'

    name = str(name).strip()
    # 保留中文字符，只替换特殊字符
    name = re.sub(r'[^\w\u4e00-\u9fff]', '_', name)
    name = re.sub(r'_+', '_', name)
    name = name.strip('_')

    if name and name[0].isdigit():
        name = 'col_' + name
    return name


def is_date_value(val):
    """判断值是否像日期"""
    if val is None:
        return False

    val_str = str(val).strip()

    # 空的或太短的不算
    if not val_str or len(val_str) < 5:
        return False

    # 常见日期格式模式
    date_patterns = [
        r'^\d{4}-\d{1,2}-\d{1,2}',           # 2024-01-15
        r'^\d{4}/\d{1,2}/\d{1,2}',             # 2024/01/15
        r'^\d{4}\.\d{1,2}\.\d{1,2}',           # 2024.01.15
        r'^\d{1,2}-\d{1,2}-\d{4}',             # 01-15-2024
        r'^\d{1,2}/\d{1,2}/\d{4}',             # 01/15/2024
        r'^\d{1,2}\.\d{1,2}\.\d{4}',           # 01.15.2024
        r'^\d{4}年\d{1,2}月\d{1,2}日',         # 2024年01月15日
        r'^\d{1,2}月\d{1,2}日',                # 01月15日
    ]

    for pattern in date_patterns:
        if re.match(pattern, val_str):
            return True

    # 检查是否包含日期关键词
    date_keywords = ['年', '月', '日', '时', '分', '秒']
    if any(kw in val_str for kw in date_keywords) and re.search(r'\d', val_str):
        return True

    return False


def is_integer_value(val):
    """判断值是否像整数"""
    if val is None:
        return False

    val_str = str(val).strip()

    # 去除常见前缀
    for prefix in ['¥', '$', '€', '£', '¥', ' ']:
        val_str = val_str.replace(prefix, '')

    # 去除逗号
    val_str = val_str.replace(',', '')

    # 空字符串不算
    if not val_str or val_str in ['-', '.', '-.']:
        return False

    # 检查是否是整数
    try:
        int(val_str)
        # 检查是否像日期（包含-且两边有数字）
        if '-' in val_str and re.match(r'\d.*-\d', val_str):
            return False
        return True
    except (ValueError, TypeError):
        return False


def is_decimal_value(val):
    """判断值是否像浮点数"""
    if val is None:
        return False

    val_str = str(val).strip()

    # 去除货币符号和逗号
    for prefix in ['¥', '$', '€', '£', ' ']:
        val_str = val_str.replace(prefix, '')

    val_str = val_str.replace(',', '')

    # 空的不算
    if not val_str or val_str in ['-', '.', '-.', '-0', '0']:
        return False

    try:
        float(val_str)
        # 确保有小数点或科学计数法
        if '.' in val_str or 'e' in val_str.lower():
            return True
        return False
    except (ValueError, TypeError):
        return False


def infer_column_type(values):
    """根据列的值推断最佳字段类型"""
    if not values:
        return 'VARCHAR(255)'

    non_empty_values = [v for v in values if v is not None and str(v).strip() != '']

    if not non_empty_values:
        return 'VARCHAR(255)'

    total = len(non_empty_values)

    # 统计各类型占比
    int_count = sum(1 for v in non_empty_values if is_integer_value(v))
    decimal_count = sum(1 for v in non_empty_values if is_decimal_value(v))
    date_count = sum(1 for v in non_empty_values if is_date_value(v))

    int_ratio = int_count / total
    decimal_ratio = decimal_count / total
    date_ratio = date_count / total

    # 如果70%以上是整数
    if int_ratio > 0.7:
        return 'INT'

    # 如果70%以上是浮点数
    if decimal_ratio > 0.7:
        return 'DECIMAL(15,4)'

    # 如果70%以上像日期
    if date_ratio > 0.7:
        return 'VARCHAR(50)'

    # 混合类型：整数+浮点数 -> DECIMAL
    if int_ratio + decimal_ratio > 0.7:
        return 'DECIMAL(15,4)'

    # 默认文本类型
    return 'VARCHAR(255)'


def create_table(connection, header, data_rows):
    """根据表头创建表"""
    cursor = connection.cursor()

    # 处理重复列名
    columns = []
    seen_names = {}

    for col in header:
        col_clean = sanitize_column_name(col)
        if col_clean in seen_names:
            seen_names[col_clean] += 1
            col_clean = f"{col_clean}_{seen_names[col_clean]}"
        else:
            seen_names[col_clean] = 0
        columns.append(col_clean)

    # 推断每列类型
    col_definitions = []
    full_columns = columns.copy()

    for i, col in enumerate(columns):
        col_values = [row[i] if i < len(row) else None for row in data_rows[:100]] if data_rows else []
        col_type = infer_column_type(col_values)
        col_definitions.append(f"`{col}` {col_type}")

    # 添加额外字段
    col_definitions.append("`source_file` VARCHAR(255)")
    col_definitions.append("`import_date` DATE")
    full_columns.extend(['source_file', 'import_date'])

    cursor.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")

    create_sql = f"""
    CREATE TABLE {TABLE_NAME} (
        id INT AUTO_INCREMENT PRIMARY KEY,
        {', '.join(col_definitions)}
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    """

    cursor.execute(create_sql)
    connection.commit()
    cursor.close()

    print(f"  表 {TABLE_NAME} 已创建，共 {len(full_columns)} 列")
    return full_columns


def convert_value(val, inferred_type):
    """根据推断的类型转换值"""
    if val is None:
        return None

    val_str = str(val).strip()

    # 空字符串转None
    if val_str == '':
        return None

    # 根据推断类型转换
    if inferred_type == 'INT':
        try:
            for prefix in ['¥', '$', '€', '£', ' ']:
                val_str = val_str.replace(prefix, '')
            val_str = val_str.replace(',', '')
            return int(float(val_str))
        except (ValueError, TypeError):
            return None

    elif inferred_type == 'DECIMAL(15,4)':
        try:
            for prefix in ['¥', '$', '€', '£', ' ']:
                val_str = val_str.replace(prefix, '')
            val_str = val_str.replace(',', '')
            return float(val_str)
        except (ValueError, TypeError):
            return None

    elif inferred_type == 'VARCHAR(50)':
        # 日期类型，直接返回字符串
        return val_str

    else:
        # VARCHAR 类型
        return val_str


def insert_data(connection, data_rows, columns, source_file, column_types):
    """插入数据"""
    if not data_rows:
        return 0

    cursor = connection.cursor()

    num_cols = len(columns)
    placeholders = ', '.join(['%s'] * num_cols)
    col_names_str = ', '.join([f'`{c}`' for c in columns])
    insert_sql = f"INSERT INTO {TABLE_NAME} ({col_names_str}) VALUES ({placeholders})"

    import_date = datetime.now().strftime('%Y-%m-%d')

    batch_size = 500
    batch = []
    inserted_count = 0

    for row in data_rows:
        # 转换每列的值
        row_list = []
        for i in range(num_cols - 2):  # 排除 source_file 和 import_date
            if i < len(row):
                row_list.append(convert_value(row[i], column_types[i]))
            else:
                row_list.append(None)

        row_data = row_list + [source_file, import_date]
        batch.append(tuple(row_data))

        if len(batch) >= batch_size:
            try:
                cursor.executemany(insert_sql, batch)
                connection.commit()
                inserted_count += len(batch)
            except Exception as e:
                print(f"    批量插入错误: {e}")
                connection.rollback()
                # 逐条尝试
                for single_row in batch:
                    try:
                        cursor.execute(insert_sql, single_row)
                        connection.commit()
                        inserted_count += 1
                    except:
                        pass
            batch = []

    # 提交剩余数据
    if batch:
        try:
            cursor.executemany(insert_sql, batch)
            connection.commit()
            inserted_count += len(batch)
        except Exception as e:
            print(f"    最后批量插入错误: {e}")
            connection.rollback()
            for single_row in batch:
                try:
                    cursor.execute(insert_sql, single_row)
                    connection.commit()
                    inserted_count += 1
                except:
                    pass

    cursor.close()
    return inserted_count


def main():
    print("=" * 60)
    print("开始导入B2C订单费用数据")
    print("=" * 60)

    print("\n1. 连接数据库...")
    try:
        connection = pymysql.connect(**DB_CONFIG)
        print("   数据库连接成功")
    except Exception as e:
        print(f"   数据库连接失败: {e}")
        return

    print("\n2. 创建数据库...")
    create_database(connection)

    xlsx_files = [f for f in os.listdir(XLSX_DIR) if f.endswith('.xlsx')]
    print(f"\n3. 扫描文件夹: {XLSX_DIR}")
    print(f"   找到 {len(xlsx_files)} 个xlsx文件")

    if not xlsx_files:
        print("   未找到任何xlsx文件")
        connection.close()
        return

    all_header = None
    all_columns = None
    all_data = []
    all_column_types = None
    file_count = 0

    for xlsx_file in sorted(xlsx_files):
        filepath = os.path.join(XLSX_DIR, xlsx_file)
        print(f"\n  处理: {xlsx_file}")

        header, data_rows = read_xlsx_sheet(filepath, SHEET_NAME)

        if not header:
            print(f"    未找到表单 '{SHEET_NAME}' 或无数据")
            continue

        if all_header is None:
            all_header = header
            all_columns = create_table(connection, header, data_rows)

            # 推断所有列的类型
            all_column_types = []
            for i in range(len(header)):
                col_values = [row[i] if i < len(row) else None for row in data_rows[:100]] if data_rows else []
                col_type = infer_column_type(col_values)
                all_column_types.append(col_type)

            print(f"    表头列数: {len(header)}")
            print(f"    字段类型推断:")
            for j, (col, col_type) in enumerate(zip(all_columns[:-2], all_column_types)):
                print(f"      {j+1}. {col[:30]:30} -> {col_type}")

        if data_rows:
            all_data.extend(data_rows)
            file_count += 1
            print(f"    读取 {len(data_rows)} 条数据")

    if not all_data:
        print("\n未找到任何数据")
        connection.close()
        return

    print(f"\n4. 合并导入 {file_count} 个文件, 共 {len(all_data)} 条数据...")

    count = insert_data(connection, all_data, all_columns, 'merged_files', all_column_types)
    print(f"    插入 {count} 条数据")

    print("\n" + "=" * 60)
    print(f"处理完成! 共导入 {count} 条数据")
    print("=" * 60)

    # 显示表结构
    cursor = connection.cursor()
    cursor.execute(f"DESCRIBE {TABLE_NAME}")
    print("\n表结构:")
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1]}")

    cursor.execute(f"SELECT COUNT(*) FROM {TABLE_NAME}")
    print(f"\n总记录数: {cursor.fetchone()[0]}")

    cursor.close()
    connection.close()


if __name__ == '__main__':
    main()
